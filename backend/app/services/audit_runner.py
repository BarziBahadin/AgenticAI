"""
app/services/audit_runner.py
Batch audit job orchestration.
Fetches chat transcripts from DB, sends to LLM, persists results as NDJSON, builds Excel.
"""
import asyncio
import logging
import re
import uuid
from typing import Optional

import aiomysql
import openpyxl

from app.config import settings
from app.services.audit_storage import (
    append_ndjson,
    init_job_files,
    ndjson_path,
    now_iso,
    read_ndjson_tail,
    read_state,
    write_state,
    xlsx_path,
)
from app.services.llm_factory import get_llm_service
from app.services.schema_detect import get_chat_columns, pick_agent_column

logger = logging.getLogger(__name__)

# Active jobs: job_id → asyncio.Task
_running: dict[str, asyncio.Task] = {}


# ── DB helpers ────────────────────────────────────────────────────────────────

async def _get_conn():
    return await aiomysql.connect(
        host=settings.db_host,
        port=settings.db_port,
        user=settings.db_user,
        password=settings.db_password,
        db=settings.db_name,
    )


async def _fetch_latest_request_ids(limit: int) -> list[int]:
    conn = await _get_conn()
    try:
        async with conn.cursor() as cur:
            await cur.execute(
                """
                SELECT br.id
                FROM base_requests br
                WHERE EXISTS (SELECT 1 FROM base_chats bc WHERE bc.request_id = br.id)
                ORDER BY br.created_at DESC, br.id DESC
                LIMIT %s
                """,
                (limit,),
            )
            rows = await cur.fetchall()
        return [int(r[0]) for r in rows]
    finally:
        conn.close()


async def _fetch_transcript(request_id: int) -> list[dict]:
    cols = await get_chat_columns()
    agent_col = pick_agent_column(cols)
    has_account_type = "account_type" in cols

    select_cols = ", ".join([
        "id",
        "request_id",
        "account_type" if has_account_type else "NULL AS account_type",
        "message",
        "sent_at",
        f"{agent_col} AS agent_ref" if agent_col else "NULL AS agent_ref",
    ])

    conn = await _get_conn()
    try:
        async with conn.cursor(aiomysql.DictCursor) as cur:
            await cur.execute(
                f"""
                SELECT {select_cols}
                FROM base_chats
                WHERE request_id = %s
                ORDER BY sent_at DESC, id DESC
                LIMIT %s
                """,
                (request_id, settings.audit_max_messages),
            )
            rows = await cur.fetchall()
    finally:
        conn.close()

    rows = list(reversed(rows))  # restore chronological order

    agent_types = {
        t.strip().lower()
        for t in settings.agent_account_types.split(",")
        if t.strip()
    }
    max_chars = settings.audit_max_chars_per_msg

    transcript = []
    for i, r in enumerate(rows):
        account_type = str(r.get("account_type") or "").lower()
        is_agent = (r.get("agent_ref") is not None) if agent_col else (account_type in agent_types)
        msg = str(r.get("message") or "")
        if len(msg) > max_chars:
            msg = msg[:max_chars] + "…"
        transcript.append({
            "index": i + 1,
            "role": "agent" if is_agent else "customer",
            "text": msg,
            "ts": r["sent_at"].isoformat() if r.get("sent_at") else "",
        })

    return transcript


# ── Job lifecycle ─────────────────────────────────────────────────────────────

async def start_audit_job() -> dict:
    if len(_running) >= settings.audit_max_concurrent_jobs:
        raise RuntimeError(
            f"Max concurrent jobs ({settings.audit_max_concurrent_jobs}) reached. "
            "Stop an existing job first."
        )

    # Verify provider is ready
    provider = settings.audit_provider.lower()
    if provider == "ollama":
        from app.services.ollama_service import ollama_service
        if not await ollama_service.health_check():
            raise RuntimeError(
                f"Ollama is not available at {settings.ollama_host}. "
                "Make sure Ollama is running and the model is loaded."
            )
    elif provider == "gemini" and not settings.gemini_api_key:
        raise RuntimeError("GEMINI_API_KEY is not set.")
    elif provider == "together" and not settings.together_api_key:
        raise RuntimeError("TOGETHER_API_KEY is not set.")

    job_id = str(uuid.uuid4())
    await asyncio.to_thread(init_job_files, job_id)

    state = {
        "job_id": job_id,
        "status": "queued",
        "provider": provider,
        "model": settings.ollama_model if provider == "ollama" else (
            settings.gemini_model if provider == "gemini" else settings.together_model
        ),
        "max_chats": settings.audit_max_chats,
        "created_at": now_iso(),
        "started_at": None,
        "finished_at": None,
        "stop_requested": False,
        "processed": 0,
        "success": 0,
        "failed": 0,
        "total_estimate": None,
        "last_error": None,
        "recent_errors": [],
        "files": {"ndjson": str(ndjson_path(job_id))},
    }
    await asyncio.to_thread(write_state, job_id, state)

    task = asyncio.create_task(_run_job(job_id))
    task.add_done_callback(lambda _: _running.pop(job_id, None))
    _running[job_id] = task

    return {"job_id": job_id, "model": state["model"], "provider": provider}


async def stop_audit_job(job_id: str) -> None:
    st = await asyncio.to_thread(read_state, job_id)
    if not st:
        raise ValueError(f"Job {job_id} not found")
    if st["status"] in ("stopped", "completed", "failed"):
        return
    st["stop_requested"] = True
    st["status"] = "stopped"
    st["finished_at"] = now_iso()
    await asyncio.to_thread(write_state, job_id, st)


async def get_audit_state(job_id: str) -> Optional[dict]:
    return await asyncio.to_thread(read_state, job_id)


def get_running_job_ids() -> list[str]:
    return list(_running.keys())


# ── Main job loop ─────────────────────────────────────────────────────────────

async def _is_stop_requested(job_id: str) -> bool:
    st = await asyncio.to_thread(read_state, job_id)
    return bool(st and (st.get("stop_requested") or st.get("status") == "stopped"))


async def _run_job(job_id: str) -> None:
    st = await asyncio.to_thread(read_state, job_id)
    if not st:
        return

    st["status"] = "running"
    st["started_at"] = now_iso()
    await asyncio.to_thread(write_state, job_id, st)

    llm = get_llm_service()
    opts = {
        "temperature": settings.audit_temperature,
        "max_tokens": settings.audit_max_tokens,
    }

    try:
        ids = await _fetch_latest_request_ids(st["max_chats"])
    except Exception as e:
        st["status"] = "failed"
        st["finished_at"] = now_iso()
        st["last_error"] = str(e)
        await asyncio.to_thread(write_state, job_id, st)
        return

    st["total_estimate"] = len(ids)
    await asyncio.to_thread(write_state, job_id, st)

    STATE_WRITE_INTERVAL = 1  # write after every item so counters stay live
    last_state_write = 0

    for request_id in ids:
        if await _is_stop_requested(job_id):
            break

        try:
            st = await asyncio.to_thread(read_state, job_id) or st
            transcript = await _fetch_transcript(request_id)

            if await _is_stop_requested(job_id):
                break

            if not transcript:
                st["processed"] += 1
                st["failed"] += 1
                await asyncio.to_thread(append_ndjson, job_id, {
                    "request_id": request_id,
                    "audited_at": now_iso(),
                    "model": st["model"],
                    "ok": False,
                    "error": "Empty transcript",
                })
                if st["processed"] - last_state_write >= STATE_WRITE_INTERVAL:
                    await asyncio.to_thread(write_state, job_id, st)
                    last_state_write = st["processed"]
                continue

            audit_input = {
                "chat_id": str(request_id),
                "language": "ar",
                "sla_thresholds": {"frt_seconds": 120, "wait_gap_seconds": 600},
                "transcript": transcript,
            }

            if await _is_stop_requested(job_id):
                break

            audit = await llm.generate_chat_audit(audit_input, opts)

            st = await asyncio.to_thread(read_state, job_id) or st
            st["processed"] += 1
            st["success"] += 1

            await asyncio.to_thread(append_ndjson, job_id, {
                "request_id": request_id,
                "audited_at": now_iso(),
                "model": st["model"],
                "ok": True,
                "audit": audit,
            })

            if st["processed"] - last_state_write >= STATE_WRITE_INTERVAL:
                await asyncio.to_thread(write_state, job_id, st)
                last_state_write = st["processed"]

        except Exception as e:
            if await _is_stop_requested(job_id):
                break

            st = await asyncio.to_thread(read_state, job_id) or st
            st["processed"] += 1
            st["failed"] += 1
            msg = str(e)
            st["last_error"] = msg
            st["recent_errors"] = ([{"at": now_iso(), "request_id": request_id, "message": msg}]
                                    + st.get("recent_errors", []))[:50]

            await asyncio.to_thread(append_ndjson, job_id, {
                "request_id": request_id,
                "audited_at": now_iso(),
                "model": st["model"],
                "ok": False,
                "error": msg,
            })
            await asyncio.to_thread(write_state, job_id, st)
            last_state_write = st["processed"]

            # Rate-limit backoff (Gemini / Together)
            wait_ms = _parse_retry_delay(msg)
            if wait_ms:
                await asyncio.sleep(wait_ms / 1000)

    # Final state flush
    st = await asyncio.to_thread(read_state, job_id) or st

    if await _is_stop_requested(job_id):
        st["status"] = "stopped"
        st["finished_at"] = now_iso()
        await asyncio.to_thread(write_state, job_id, st)
        return

    # Build Excel
    try:
        await asyncio.to_thread(_build_excel, job_id)
        st = await asyncio.to_thread(read_state, job_id) or st
        st["status"] = "completed"
        st["finished_at"] = now_iso()
        st["files"]["xlsx"] = str(xlsx_path(job_id))
    except Exception as e:
        logger.error(f"Excel build failed for {job_id}: {e}")
        st["status"] = "completed"
        st["finished_at"] = now_iso()
        st["last_error"] = f"Excel build failed: {e}"

    await asyncio.to_thread(write_state, job_id, st)


def _parse_retry_delay(msg: str) -> Optional[int]:
    """Extract retry wait in ms from rate-limit error messages."""
    m = re.search(r'"retryDelay"\s*:\s*"(\d+)s"', msg)
    if m:
        return (int(m.group(1)) + 1) * 1000
    m = re.search(r'retry in\s+([\d.]+)s', msg, re.I)
    if m:
        return (int(float(m.group(1))) + 1) * 1000
    if "429" in msg or "quota" in msg.lower() or "resource_exhausted" in msg.lower():
        return 30_000
    return None


# ── Excel export ──────────────────────────────────────────────────────────────

def _build_excel(job_id: str) -> None:
    import json as _json

    out = xlsx_path(job_id)
    st = read_state(job_id)

    wb = openpyxl.Workbook(write_only=True)

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["job_id", job_id])
    if st:
        ws_sum.append(["provider", st.get("provider", "")])
        ws_sum.append(["model", st.get("model", "")])
        ws_sum.append(["status", st.get("status", "")])
        ws_sum.append(["created_at", st.get("created_at", "")])
        ws_sum.append(["started_at", st.get("started_at", "")])
        ws_sum.append(["finished_at", st.get("finished_at", "")])
        ws_sum.append(["processed", st.get("processed", 0)])
        ws_sum.append(["success", st.get("success", 0)])
        ws_sum.append(["failed", st.get("failed", 0)])

    # Audits sheet
    ws = wb.create_sheet("Audits")
    ws.append([
        "request_id", "audited_at", "model", "ok",
        "score_total", "score_compliance", "score_quality", "score_resolution", "score_sla",
        "risk_level", "sentiment", "category",
        "summary", "checks_count", "coaching_count", "error",
    ])

    ndjson = ndjson_path(job_id)
    if ndjson.exists():
        with open(ndjson, encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if not s:
                    continue
                try:
                    obj = _json.loads(s)
                except Exception:
                    continue
                audit = obj.get("audit") or {}
                scores = audit.get("scores") or {}
                ws.append([
                    str(obj.get("request_id", "")),
                    str(obj.get("audited_at", "")),
                    str(obj.get("model", "")),
                    "true" if obj.get("ok") else "false",
                    scores.get("total", ""),
                    scores.get("compliance", ""),
                    scores.get("quality", ""),
                    scores.get("resolution", ""),
                    scores.get("sla", ""),
                    str(audit.get("risk_level", "")),
                    str(audit.get("sentiment", "")),
                    str(audit.get("category", "")),
                    str(audit.get("summary", ""))[:500],
                    len(audit.get("checks") or []),
                    len(audit.get("coaching") or []),
                    str(obj.get("error", ""))[:500],
                ])

    wb.save(out)
    logger.info(f"Excel saved: {out}")
